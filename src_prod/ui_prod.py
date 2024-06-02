import tkinter as tk
from tkinter import ttk, Toplevel
import openpyxl
import pandas as pd 
import numpy as np
import os
import optimization_service_prod
from mixing_service_prod import Mixing_combined_solution
from OT_service_prod import OT_service

class UI:
    def __init__(self, root):
        self._root = root
        self.service = optimization_service_prod.OptimizationService()
        self.model = None
        self.batch_labels = []
        self.batch_entries = []
        self.chosen_theme = 'forest-light'
        self.min_entry = None
        self.max_entry = None
        self.cat_values_entry = None
        self.path = None
        #to be defined in config
        self.data_dir = "data"
        #to be defined in  read parameters window/functions
        self.ot = True
        self.geometry_x = 1000
        self.geometry_y = 500
        self.options_run = False
        self.model_build = False



    def start_options(self):
        
        if not self.options_run:
            self.style = ttk.Style(self._root)
            self._root.tk.call("source", "forest-light.tcl")
            self.options_run = True
            self.style.theme_use(self.chosen_theme)
        else:
            self.frame.destroy()

        self._root.title("Settings")
        self._root.geometry(str(self.geometry_x) + "x" + str(self.geometry_y))
        self.frame = ttk.Frame(self._root, height=self.geometry_y, width=self.geometry_x)
        self.frame.pack()
        self.widget_window('Settings')

        self.file_name_entry = ttk.Entry(self.widgets_frame)
        self.file_name_entry.insert(0, "settings.xlsx")
        self.file_name_entry.bind("<FocusIn>", lambda e: self.user_name_entry.delete('0', 'end'))
        self.file_name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="e")

        self.load_data_button = ttk.Button(self.widgets_frame, text="Load data", command=self.load_data)
        self.load_data_button.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="ew")

        
        self.load_data()
        self.build_treeview()
        self.load_data_to_treeview()
        

        delete_button = ttk.Button(self.widgets_frame, text="Delete", command=self.delete_row)
        delete_button.grid(row=2, column=0, padx=5, pady=5, sticky="nsew")

        self.toggle_mode_switch = ttk.Checkbutton(self.widgets_frame, text="Dark mode", style="Switch", command=self.toggle_mode)
        self.toggle_mode_switch.grid(row=3, column=0, padx=5, pady=10, sticky="nsew")

        self.next_button = ttk.Button(self.treeFrame, text="Optimization", command=self.initiate_settings)
        self.next_button.pack(side="right", padx=5, pady=5)
        
        self.ot_button = ttk.Button(self.treeFrame, text="OT", command=self._handle_start_ot_options_button_click)
        self.ot_button.pack(side="right", padx=5, pady=5)

    def _handle_start_ot_options_button_click(self):
        self.ot = True
        self.start_ot_options()


    def start_parameters(self):
        self.frame.destroy()
        #self.settings = list(self.sheet.values)
        self.style = ttk.Style(self._root)
        self.style.theme_use(self.chosen_theme)
        self._root.title("Initialize parameters")
        
        self._root.geometry(str(self.geometry_x) + "x" + str(self.geometry_y))
        self.frame = ttk.Frame(self._root, height=self.geometry_y, width=self.geometry_x)
        self.frame.pack()
        #self.frame.geometry("550x450")
        self.widget_window('set parameters')
        
        self.file_name_entry = ttk.Entry(self.widgets_frame)
        self.file_name_entry.insert(0, self.reaction_parameter_file_name)
        self.file_name_entry.bind("<FocusIn>", lambda e: self.name_entry.delete('0', 'end'))
        self.file_name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

        self.load_data_button = ttk.Button(self.widgets_frame, text="Load data", command=self.load_data)
        self.load_data_button.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="ew")

        delete_button = ttk.Button(self.widgets_frame, text="Delete", command=self.delete_row)
        delete_button.grid(row=2, column=0, padx=5, pady=5, sticky="nsew")

        separator = ttk.Separator(self.widgets_frame)
        separator.grid(row=3, column=0, padx=(20, 10), pady=10, sticky="ew")
        self.load_data()
        self.build_treeview()
        self.load_data_to_treeview()
        self.back_button = ttk.Button(self.treeFrame, text="Back", command=self.start_options)
        self.back_button.pack(side="left", padx=5, pady=5)
        self.next_button = ttk.Button(self.treeFrame, text="Next", command=self.init_experiment)
        self.next_button.pack(side="right", padx=5, pady=5)
        self.simulate_button = ttk.Button(self.treeFrame, text="Simulate", command=self._handle_simulate_with_synthetic_data_button_click)
        self.simulate_button.pack(side="right", padx=5, pady=5)

    def _handle_simulate_with_synthetic_data_button_click(self):
        self.service.create_params(list(self.sheet.values))
        exp_name = self.return_value_by_key(self.settings, 'experiment name')[0]
        self.service.create_experiment(exp_name, minimize=True)
        if self.simulate_one_or_all:
            self.simulate_all()
        else:
            self.simulate()

    def init_experiment(self):
        self.service.create_params(list(self.sheet.values))
        self.service.create_experiment(self.experiment_name, minimize = False)
        self.start_experiment()

    def start_experiment(self):
        self.frame.destroy()
        self.style = ttk.Style(self._root)
        self.style.theme_use(self.chosen_theme)
        self._root.title("Initialize experiment")
        
        self._root.geometry(str(self.geometry_x) + "x" + str(self.geometry_y))
        self.frame = ttk.Frame(self._root, height=self.geometry_y, width=self.geometry_x)
        self.frame.pack()

        self.widget_window('Initialize experiment')

        
        self.file_name_entry = ttk.Entry(self.widgets_frame)
        self.file_name_entry.insert(0, self.experiment_load_file_name)
        #self.file_name_entry.bind("<FocusIn>", lambda e: self.file_name_entry.delete('0', 'end'))
        self.file_name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

        self.load_data_button = ttk.Button(self.widgets_frame, text="Load data", command=self._handle_load_data_button)
        self.load_data_button.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="ew")
        
        self.initialize_model_with_excel = ttk.Button(self.widgets_frame, text="Build model", command=self.build_model_with_initial_data)
        self.initialize_model_with_excel.grid(row=2, column=0, padx=5, pady=(0, 5), sticky="ew")

        self.generate_batch_button = ttk.Button(self.widgets_frame, text="Generate batch", command=self.generate_batch)
        self.generate_batch_button.grid(row=3, column=0, padx=5, pady=(0, 5), sticky="ew")
                
        self.arm_name_entry = ttk.Entry(self.widgets_frame)
        self.arm_name_entry.grid(row=4, column=0, padx=5, pady=(0, 5), sticky="ew")
        self.param_names = self.service.return_params_values('name')
        self.param_value_entries = []

        for i, param_name in enumerate(self.param_names):
            param_value_entry = ttk.Entry(self.widgets_frame)
            self.param_value_entries.append(param_value_entry)
            self.param_value_entries[i].insert(0, param_name)
            self.param_value_entries[i].grid(row=i+5, column=0, padx=5, pady=(0, 5), sticky="ew")

        insert_button = ttk.Button(self.widgets_frame, text="Insert", command=self.insert_experiment_row)
        insert_button.grid(row=8, column=0, padx=5, pady=5, sticky="nsew")

        delete_button = ttk.Button(self.widgets_frame, text="Delete", command=self.delete_row)
        delete_button.grid(row=9, column=0, padx=5, pady=5, sticky="nsew")
        self.cols = ["arm_name"]+self.param_names
        self.load_data()
        self.build_treeview()
        self.load_data_to_treeview()
        self.exp_num  = len(list(self.sheet.values))-1
        self.next_button = ttk.Button(self.treeFrame, text="Next", command=self.prepare_mixing)
        self.next_button.pack(side="right", padx=5, pady=5)
        self.arm_name_entry.insert(0, f"0_{self.exp_num}")
        self.render_button = ttk.Button(self.widgets_frame, text="Render", command=self._handle_render_button_click)
        self.render_button.grid(row=10, column=0, padx=5, pady=5, sticky="nsew")

        separator = ttk.Separator(self.widgets_frame)
        separator.grid(row=11, column=0, padx=(20, 10), pady=10, sticky="ew")
        self.read_results_button = ttk.Button(
          master=self.widgets_frame,
          text="Read results",
          command=self._read_results
        )
        self.read_results_button.grid(row=12, column=0, padx=5, pady=5, sticky="nsew")
        self.back_button = ttk.Button(self.treeFrame, text="Back", command=self.start_parameters)
        self.back_button.pack(side="left", padx=5, pady=5)


    def prepare_mixing(self):
        if not self.model_build:
            self.build_model_with_initial_data()
        self.service.create_reaction(self.solution_volume, self.analysis_volume)
        self.arms = self.service.return_arms()
        self.mixing_service = Mixing_combined_solution(self.service.reaction, self.arms, self.mother_solutions)
        self.save_solution_data()
        if self.ot:
            self.solutions_path = os.path.join(self.data_dir, self.solutions_load_file_name)
        self.mix_solutions()
    
    def read_settings(self):
        path = os.path.join(self.data_dir, 'settings.xlsx')
        self.settings = pd.read_excel(path)
        self.analysis_volume = self.settings[self.settings['setting']=='analysis volume']['setting value'].values[0]
        self.solution_volume = self.settings[self.settings['setting']=='solution volume']['setting value'].values[0]
        self.batch_size = self.settings[self.settings['setting']=='batch size']['setting value'].values[0]
        self.model_name = self.settings[self.settings['setting']=='model name']['setting value'].values[0]
        self.fixed_temperature = self.settings[self.settings['setting']=='fixed temperature']['setting value'].values[0]
        print(self.fixed_temperature)
        if 'yes' in self.fixed_temperature:
            self.fixed_temperature = True
        else:
            self.fixed_temperature = False
        print(self.fixed_temperature)
        self.analysis_dilution_factor = self.settings[self.settings['setting']=='analysis dilution factor']['setting value'].values[0]
        self.experiment_save_file_name = self.settings[self.settings['setting']=='experiment save file name']['setting value'].values[0]
        self.experiment_load_file_name = self.settings[self.settings['setting']=='experiment load file name']['setting value'].values[0]
        self.solutions_load_file_name = self.settings[self.settings['setting']=='solutions load file name']['setting value'].values[0]
        self.solutions_save_file_name = self.settings[self.settings['setting']=='solutions save file name']['setting value'].values[0]
        self.reaction_parameter_file_name = self.settings[self.settings['setting']=='reaction parameter file name']['setting value'].values[0]
        self.mother_solutions = self.settings[self.settings['setting']=='mother solutions']['setting value'].values[0]
        if 'yes' in self.mother_solutions:
            self.mother_solutions = True
        else:
            self.mother_solutions = False
        self.simulation_batch_size = self.settings[self.settings['setting']=='simulation batch size']['setting value'].values[0]
        self.simulation_batches = self.settings[self.settings['setting']=='simulation batches']['setting value'].values[0]
        self.simulate_one_or_all = self.settings[self.settings['setting']=='simulate all']['setting value'].values[0]
        self.results_file_name = self.settings[self.settings['setting']=='results file name']['setting value'].values[0]
        self.experiment_name = self.settings[self.settings['setting']=='experiment name']['setting value'].values[0]

    
    def mix_solutions(self):
        self.frame.destroy()
        self.style = ttk.Style(self._root)
        self.style.theme_use(self.chosen_theme)
        self._root.title("Solutions")
        self.frame = ttk.Frame(self._root, height=450, width=1000)
        self.frame.pack()
        
        self.widget_window('Mix solutions')
        self.file_name_entry = ttk.Entry(self.widgets_frame)
        self.file_name_entry.insert(0, self.solutions_load_file_name)
        self.file_name_entry.bind("<FocusIn>", lambda e: self.file_name_entry.delete('0', 'end'))
        self.file_name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

        self.load_data_button = ttk.Button(self.widgets_frame, text="Load data", command=self.load_data)
        self.load_data_button.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="ew")
        
        delete_button = ttk.Button(self.widgets_frame, text="Delete", command=self.delete_row)
        delete_button.grid(row=10, column=0, padx=5, pady=5, sticky="nsew")

        separator = ttk.Separator(self.widgets_frame)
        separator.grid(row=11, column=0, padx=(20, 10), pady=10, sticky="ew")

        self.cols = ["experiment_name", "component", "amount", "mass", "volume", "mix_order"]
        self.load_data()
        self.build_treeview()
        self.load_data_to_treeview()
        self.exp_num  = len(list(self.sheet.values))-1
        self.next_button = ttk.Button(self.treeFrame, text="Next", command=self.run_experiment)
        self.next_button.pack(side="right", padx=5, pady=5)        
        self.back_button = ttk.Button(self.treeFrame, text="Back", command=self.start_experiment)
        self.back_button.pack(side="left", padx=5, pady=5)


    def widget_window(self, text):
        self.widgets_frame = ttk.LabelFrame(self.frame, text=text, height=self.geometry_y-30, width=170)
        self.widgets_frame.grid_propagate(0)
        self.widgets_frame.grid(row=0, column=0, padx=10, pady=10)

    def return_value_by_key(self, sheet, key):
        for i, row in enumerate(sheet):
            if row[0] == key:
                return row[1:]
            
    def _handle_load_data_button(self):
        self.path = os.path.join(self.data_dir, self.file_name_entry.get())
        self.load_data()
        self.build_treeview()
        self.load_data_to_treeview()
        self.next_button.pack(side="right", padx=5, pady=5)


        

    def generate_batch(self):
        if self.service.axmodel.model:
            print('generating batch with batch size', self.batch_size)
            self.service.new_batch(self.batch_size, self.fixed_temperature)
        else:
            self.service.build_model_with_random_points(self, 'something', self.batch_size, self.fixed_temperature)

        self.path = os.path.join(self.data_dir, self.experiment_save_file_name)
        self.service.save_experiment(self.path)

    def build_model_with_initial_data(self):
        self.path = os.path.join(self.data_dir, self.file_name_entry.get())
        self.model_build = True
        print('building model with data:', self.path)
        self.service.build_model_with_initial_data(self.service.params, self.path) 

    def run_experiment(self):
        if self.ot:
            self.start_ot_options()
        else:
            self.results()

    def start_ot_options(self):
        self.frame.destroy()
        self.style = ttk.Style(self._root)
        self.style.theme_use(self.chosen_theme)
        self._root.title("Open Trons Settings")
        self.frame = ttk.Frame(self._root, height=450, width=1000)
        self.frame.pack()

        self.widget_window('insert settings')

        self.file_name_entry = ttk.Entry(self.widgets_frame)
        self.file_name_entry.insert(0, "ot_settings.xlsx")
        self.file_name_entry.bind("<FocusIn>", lambda e: self.user_name_entry.delete('0', 'end'))
        self.file_name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="e")

        self.load_data_button = ttk.Button(self.widgets_frame, text="Load data", command=self.load_data)
        self.load_data_button.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="ew")

        #self.cols = ['setting', 'setting value']        
        self.load_data()
        self.build_treeview()
        self.load_data_to_treeview()       

        delete_button = ttk.Button(self.widgets_frame, text="Delete", command=self.delete_row)
        delete_button.grid(row=9, column=0, padx=5, pady=5, sticky="nsew")

        self.toggle_mode_switch = ttk.Checkbutton(self.widgets_frame, text="Dark mode", style="Switch", command=self.toggle_mode)
        self.toggle_mode_switch.grid(row=10, column=0, padx=5, pady=10, sticky="nsew")

        self.next_button = ttk.Button(self.treeFrame, text="Next", command=self.initiate_open_trons)
        self.next_button.pack(side="right", padx=5, pady=5)
    
        self.back_button = ttk.Button(self.treeFrame, text="Back", command=self.mix_solutions)
        self.back_button.pack(side="left", padx=5, pady=5)


    def initiate_open_trons(self):
        self.ot_service = OT_service(os.path.join(self.data_dir, 'ot_settings.xlsx'), self.data_dir)
        self.ot_service.create_main_logic()
        if self.ot:
            self.start_options()
        else:
            self.start_experiment()

    def open_trons_screen(self):
        self.frame.destroy()
        self.style = ttk.Style(self._root)
        self.style.theme_use(self.chosen_theme)
        self._root.title("OpenTrons")
        
        self.frame = ttk.Frame(self._root, height=450, width=1000)
        self.frame.pack()

        self.widget_window('Initialize OT protocol')
        
        self.file_name_entry = ttk.Entry(self.widgets_frame)
        self.file_name_entry.insert(0, "ot_script.xlsx")
        self.file_name_entry.bind("<FocusIn>", lambda e: self.file_name_entry.delete('0', 'end'))
        self.file_name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

        self.load_data_button = ttk.Button(self.widgets_frame, text="Load data", command=self.load_data)
        self.load_data_button.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="ew")

        self.function_combobox = ttk.Combobox(self.widgets_frame, values=["transfer", "mix", "wait", "set temperature", "set shaker"])
        self.function_combobox.grid(row=2, column=0, padx=5, pady=5,  sticky="ew")
        params = ['source', 'destination', 'volume', 'blow_out', 'blowout_location', 'air_gap']
        
        self.enter_function_button = ttk.Button(self.widgets_frame, text="Enter function", command=self.enter_function)
        self.enter_function_button.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")
        self.enter_function_button.bind("<<ComboboxSelected>>", self.enter_function)

        self.load_data()
        self.build_treeview()
        self.load_data_to_treeview()
    
    def enter_function(self):
        function_name = self.function_combobox.get()
        params = [' ', ' ']
        if function_name == "transfer":
            params = ['source', 'destination', 'volume', 'blow_out', 'blowout_location', 'air_gap']
        self.param_combobox = ttk.Combobox(self.widgets_frame, values=params)
        self.param_combobox.grid(row=4, column=0, padx=5, pady=5,  sticky="ew")
        #self.param_combobox.bind("<<ComboboxSelected>>", self.enter_parameter())
        self.enter_param_button = ttk.Button(self.widgets_frame, text="Enter parameter", command=self.enter_parameter)
        self.enter_param_button.grid(row=5, column=0, padx=5, pady=5, sticky="nsew")
        #function_combobox.delete(0, "end")
        self.row_values = [function_name]

    def enter_parameter(self):
        param_name = self.param_combobox.get()
        param_values=[]
        if param_name=='source':
            racks = ['solvent_rack','solution_rack']
            param_values = self.create_rack_position_combos(racks)
        self.param_value_combobox = ttk.Combobox(self.widgets_frame, values=param_values)
        self.param_value_combobox.grid(row=6, column=0, padx=5, pady=(0, 5), sticky="ew")
        self.enter_command_button = ttk.Button(self.widgets_frame, text="Enter command", command=self.insert_ot_row)
        self.enter_command_button.grid(row=7, column=0, padx=5, pady=5, sticky="nsew")
        
    def create_rack_position_combos(self, racks):
        combinations = []
        for rack in racks:
            for i in range(24):
                combinations.append(f"{rack}[{self.convert_index_to_OT_location(i)}]")
        return combinations

    def enter_parameter(self):
        param_name = self.param_combobox.get()
        param_values=[]
        if param_name=='source':
            racks = ['solvent_rack','solution_rack']
            param_values = self.create_rack_position_combos(racks)
        self.param_value_combobox = ttk.Combobox(self.widgets_frame, values=param_values)
        self.param_value_combobox.grid(row=6, column=0, padx=5, pady=(0, 5), sticky="ew")
        self.enter_command_button = ttk.Button(self.widgets_frame, text="Enter command", command=self.insert_ot_row)
        self.enter_command_button.grid(row=7, column=0, padx=5, pady=5, sticky="nsew")
        
    def create_rack_position_combos(self, racks):
        combinations = []
        for rack in racks:
            for i in range(24):
                combinations.append(f"{rack}[{self.convert_index_to_OT_location(i)}]")
        return combinations
    
    def _read_results(self):
        self.opentrons_window = Toplevel(self._root)
        self.opentrons_window.title("Reading results")
        self._read_file(i=0)
        self.initialize_read_window()
    


    def _read_file(self, i):
        arms = self.service.return_arms()
        arm_name = arms[i].name
        filename = self.results_file_name.replace('_x', "_"+arm_name)
        #file = open('015.csv')
        path = os.path.join(self.data_dir, 'GC results', filename)
        file = open(path)
        for i, line in enumerate(file):
            if i==2:
                with open('corrected.csv', 'w') as f:
                    f.write(line)
            if i>2:
                with open('corrected.csv', 'a') as f:
                    f.write(line)
            
        df = pd.read_csv('corrected.csv',index_col=False)
        self.df = df

    def initialize_read_window(self):
        for i in range(self.df.shape[0]):
            try:
                peak_label = ttk.Label(master=self.opentrons_window, text=f"Peak number {i+1}: peak: {self.df['Center X'][i+1]}, area: {int(self.df['Area'][i+1])}")
                peak_label.grid(row=i, column=0, sticky="w", padx=5, pady=5)
            except:
                break
            
        is_range_label = ttk.Label(master=self.opentrons_window, text="IS peak: ")
        a_range_label = ttk.Label(master=self.opentrons_window, text="A peak: ")
        self.IS_peak_entry = ttk.Entry(master=self.opentrons_window, width=5)
        self.A_peak_entry = ttk.Entry(master=self.opentrons_window, width=5)
        enter_button = ttk.Button(
            master=self.opentrons_window,
            text="Enter",
            command=self._enter_ranges
            )
        read_all_button = ttk.Button(
            master=self.opentrons_window,
            text="Read all",
            command=self.read_all_peaks
            )
        read_all_button.grid(row=self.df.shape[0]+4, column=0, columnspan=2, sticky="ew", padx=5, pady=5)   
        is_range_label.grid(row=self.df.shape[0]+1, column=0, sticky="e", padx=5, pady=5)
        a_range_label.grid(row=self.df.shape[0]+2, column=0, sticky="e", padx=5, pady=5)
        self.IS_peak_entry.grid(row=self.df.shape[0]+1, column=1, sticky="w", padx=5, pady=5)
        self.A_peak_entry.grid(row=self.df.shape[0]+2, column=1, sticky="e", padx=5, pady=5)
        enter_button.grid(row=self.df.shape[0]+3, column=0, columnspan=2, sticky="ew", padx=5, pady=5)
    
    def read_all_peaks(self):
        data = {}
        marginal = 0.1
        current_peak = 0
        #for files
        for i, arm in enumerate(self.service.return_arms()):
            self._read_file(i)
            if i == 0:
                #create base data from first file
                for j in range(self.df.shape[0]):
                    data['peak '+str(j+1)] = [self.df['Center X'][j]]
                    data['peak '+str(j+1)+' area'] = [self.df['Area'][j]]
                data = pd.DataFrame(data)
                current_peak = self.df.shape[0]

            else:
                #for each peak in the new file
                for j in range(self.df.shape[0]):
                    x = self.df['Center X'][j]
                    #for each peak in the current data
                    for k in range(data.shape[0]):
                        new_data={}
                        # if peak exists in the current data then append the area
                        if np.abs(x-data['peak '+str(k+1)][k])<marginal:
                            new_data['peak '+str(k+1)] = [x]
                            new_data['peak '+str(k+1)+' area'] = [self.df['Area'][j]]
                            
                        else:
                            new_data['peak '+str(current_peak+1)] = ['']
                            new_data['peak '+str(current_peak+1)+' area'] = ['']
                    new_data = pd.DataFrame(new_data)
                    data = pd.concat([data, new_data], axis=1)
                
                

    def _enter_ranges(self):
        range_wideness = 0.1
        is_retention_time = float(self.df['Center X'][int(self.IS_peak_entry.get())])
        a_retention_time = float(self.df['Center X'][int(self.A_peak_entry.get())])
        is_peak_min = float(is_retention_time)-range_wideness
        is_peak_max = float(is_retention_time)+range_wideness
        a_peak_min = float(a_retention_time)-range_wideness
        a_peak_max = float(a_retention_time)+range_wideness
        is_areas = []
        a_areas = []
        for i, arm in enumerate(self.service.return_arms()):
            if i>0:
                self._read_file(i)
            is_areas.append(self._import_area(i, is_peak_min, is_peak_max))
            a_areas.append(self._import_area(i, a_peak_min, a_peak_max))
        print('is_areas: ', is_areas)
        print('a_areas: ', a_areas)
        A_A_per_A_IS = np.array(a_areas)/np.array(is_areas)
        A_A_per_A_IS= A_A_per_A_IS
        calibration_factor = 1
        concentrations = A_A_per_A_IS/calibration_factor
        #self.mixing_service.combined_solutions->None
        yields = self.service.calculate_yields(concentrations, None)
        yields = yields.tolist()
        self.service.update_model_with_batch_results(yields)

        #append the first area
    def _import_area(self, i,peak_min,peak_max):
        area = None
        for row in self.df.iterrows():
            if row[1]['Center X']>=peak_min and row[1]['Center X']<=peak_max:
                area = row[1]['Area']
        return area
        
    def initiate_settings(self):
        self.read_settings()
        self.start_parameters()

    def convert_index_to_OT_location(self, index):
        if index < 6:
            return 'A' + str(index+1)
        elif index < 12:
            return 'B' + str(index-5)
        elif index < 18:
            return 'C' + str(index-11)
        elif index < 24:
            return 'D' + str(index-17)

    def insert_ot_row(self):
        row_values = []
        function_name = self.function_combobox.get()
        row_values.append(function_name)
        param_name = self.param_combobox.get()
        param_value = self.param_value_combobox.get()
        row_values.append(param_name+"="+param_value)
        # Insert row into Excel sheet
        path = self.path
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        sheet.append(row_values)
        workbook.save(path)
        
        # Insert row into treeview
        self.treeview.insert('', tk.END, values=row_values)

    def remove_none(self, data):
        # Clear existing items in the treeview
        updated = []
        for item in data:
            item_to_updated = []
            for value in item:
                #if none:
                if value == None:
                    value = ""
                item_to_updated.append(value)
            updated.append(item_to_updated)
        return updated
            
    def build_treeview(self):
        self.treeFrame = ttk.Frame(self.frame, height=450, width=810)
        self.treeFrame.grid(row=0, column=1, pady=0, padx=0, sticky="e")
        self.treeFrame.propagate(0)
        treeScroll = ttk.Scrollbar(self.treeFrame)
        treeScroll.pack(side="right", fill="y")
        
        self.treeview = ttk.Treeview(self.treeFrame, show="headings",
                                yscrollcommand=treeScroll.set, columns=self.cols, height=13)
        
        total_width = 790
        total_chars = 0
        for col in self.cols:
            total_chars += len(col)
        size_factor = int(total_width/total_chars)
        for col in self.cols:
            self.treeview.column(col, width=len(col)*size_factor)

        self.treeview.pack()
        treeScroll.config(command=self.treeview.yview)

    def browse_file(self):
        pass
    
    def insert_experiment_row(self):
        row_values = []
        name = self.arm_name_entry.get()
        row_values.append(name)
        self.arm_name_entry.delete(0, "end")
        self.arm_name_entry.insert(0, "arm name")

        for i, param_entry in enumerate(self.param_value_entries):
            row_values.append(param_entry.get())
            self.param_value_entries[i].delete(0, "end")
            self.param_value_entries[i].insert(0, self.param_names[i])
        # Insert row into Excel sheet
        path = self.path
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        sheet.append(row_values)
        workbook.save(path)
        # Insert row into treeview
        self.treeview.insert('', tk.END, values=row_values)
    def insert_row(self):
        pass

    def toggle_mode(self):
        if self.mode_switch.instate(["selected"]):
            self.style.theme_use("forest-light")
        else:
            self.style.theme_use("forest-dark")

    def delete_row(self):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.active
        list_values = list(self.sheet.values)
        n_deleted = 0

        for deleted_name in self.treeview.selection():
            sheet.delete_rows(int(deleted_name[-1])+1-n_deleted)
            n_deleted += 1 

        workbook.save(self.path)
        self.treeview.delete(self.treeview.selection())
        
    def load_data(self):
            filename = self.file_name_entry.get()
            print('filename ', filename )
            print('path ', self.data_dir)
            self.path = os.path.join(self.data_dir, filename)
            workbook = openpyxl.load_workbook(self.path)
            self.sheet = workbook.active
            list_values = list(self.sheet.values)
            self.cols = list_values[0]


    def load_data_to_treeview(self):
            list_values = list(self.sheet.values)
            list_values = self.remove_none(list_values)

            for col_name in self.cols:
                self.treeview.heading(col_name, text=col_name)

            for value_tuple in list_values[1:]:
                formatted_tuple = []
                for value in value_tuple:
                    formatted_tuple.append(self.format_number(value))
                self.treeview.insert('', tk.END, values=formatted_tuple)

    def _handle_render_button_click(self):
        self.service.render_all()

    def simulate_all(self):
        print('simulate all')
        simulation_batch_sizes = [2,4,6]
        simulation_batches = 10
        #functions = ['qPI', 'qEI',  'qUCB', 'qlogEI']
        functions = ['qlogEI', 'qEI']
        for batch_size in simulation_batch_sizes:
            for function in functions:
                save_path = os.path.join(self.data_dir, 'simulation_results_'+function+'_'+str(batch_size)+'.xlsx')
                self.service.synthetic_data_algorithm(n_batch = simulation_batches, batch_size = batch_size, function = function, save_path = save_path)
                exp_name = self.return_value_by_key(self.settings, 'experiment name')[0]
                print('experement name', exp_name)
                self.service.axmodel.model = None
                self.service.create_experiment(exp_name, minimize=True)

    def simulate(self):
        print('simulate one')
        self.service.synthetic_data_algorithm(n_batch = self.simulation_batches, batch_size = self.simulation_batch_size, function = 'qlogEI', save_path=os.path.join(self.data_dir, 'simulation_results.xlsx'))

    def trigger_min_max_values(self):
        if not self.type2_switch.instate(["selected"]):
            if self.cat_values_entry:
                self.cat_values_entry.destroy()
            self.min_entry = ttk.Entry(self.widgets_frame)
            self.min_entry.insert(0, "Parameter min value")
            self.min_entry.bind("<FocusIn>", lambda e: self.min_entry.delete('0', 'end'))
            self.min_entry.grid(row=4, column=0, padx=5, pady=(0, 5), sticky="ew")

            self.max_entry = ttk.Entry(self.widgets_frame)
            self.max_entry.insert(0, "Parameter max value")
            self.max_entry.bind("<FocusIn>", lambda e: self.max_entry.delete('0', 'end'))
            self.max_entry.grid(row=5, column=0, padx=5, pady=(0, 5), sticky="ew")

        else:
            self.min_entry.destroy()
            self.max_entry.destroy()
            self.cat_values_entry = ttk.Entry(self.widgets_frame)
            self.cat_values_entry.insert(0, "Parameter values")
            self.cat_values_entry.bind("<FocusIn>", lambda e: self.cat_values_entry.delete('0', 'end'))
            self.cat_values_entry.grid(row=4, column=0, padx=5, pady=(0, 5), sticky="ew")

    def insert_parameter_row(self):
        name = self.name_entry.get()
        parameter_type = self.type_combobox.get()
        if self.type2_switch.instate(["selected"]):
            parameter_type2 = "Categorical"
            cat_values = self.cat_values_entry.get()
            min_value = None
            max_value = None
        else:
            parameter_type2 = "Continuous"
            cat_values = None
            min_value = float(self.min_entry.get())
            max_value = float(self.max_entry.get())
        
        unit = None
        # Insert row into Excel sheet
        path = self.path
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        row_values = [name, parameter_type, parameter_type2, min_value, max_value, cat_values, unit]
        sheet.append(row_values)
        workbook.save(path)


        # Insert row into treeview
        self.treeview.insert('', tk.END, values=row_values)

        # Clear the values
        self.name_entry.delete(0, "end")
        self.name_entry.insert(0, "Name")
    
    def format_number(self, value):
        if type(value)==float:
            return f'{value:.2f}'
        else :
            return value

    def save_solution_data(self):
        data = {
    'experiment name': [],
    'component': [],
    'name': [],
    'amount': [],
    'mass': [],
    'volume': [],
    'mix order': [],
    'ms-factor': [],
    'temperature': [],
    'shaker speed': [],
    'time': []
}
        ms_solutions={
        'experiment name': [],
        'component': [],
        'name': [],
        'volume':[],
        'mix order': []
        }
        df = pd.DataFrame()
        for i, combined_solution in enumerate(self.mixing_service.combined_solutions):
            #General reaction info
            data['experiment name'].append(f'Reaction {self.arms[i].name}')
            data['component'].append('reaction')
            data['name'].append('')
            data['amount'].append('')
            data['mass'].append('')
            data['volume'].append('')
            data['mix order'].append('')
            data['ms-factor'].append('')
            data['temperature'].append(combined_solution.temperature)
            data['shaker speed'].append(combined_solution.shaker_speed)
            data['time'].append(combined_solution.time)


            #Solvent
            data['experiment name'].append(f'Reaction {self.arms[i].name}')
            data['component'].append(self.mixing_service.reaction.solvent.component)
            data['name'].append(self.mixing_service.reaction.solvent.name)
            data['amount'].append('')
            data['mass'].append('')
            data['volume'].append(combined_solution.solvent_volume)
            data['mix order'].append(1)
            data['ms-factor'].append('')
            data['temperature'].append('')
            data['shaker speed'].append('')
            data['time'].append('')


            #base/acids
            for j, base_acid in enumerate(self.mixing_service.reaction.base_acids):
                data['experiment name'].append(f'Reaction {self.arms[i].name}')
                data['name'].append(base_acid.name)
                data['component'].append(base_acid.component)
                data['amount'].append(combined_solution.base_acid_amounts[j])
                data['mass'].append(combined_solution.base_acid_masses[j])
                if 'solid' in base_acid.name:
                    data['volume'].append('')
                    data['mix order'].append('')
                else:
                    data['volume'].append(combined_solution.base_acid_volumes[j])
                    data['mix order'].append(2+j)
                data['ms-factor'].append('')
                data['temperature'].append('')
                data['shaker speed'].append('')
                data['time'].append('')

            #reactants
            for j, solution in enumerate(combined_solution.solutions):
                data['experiment name'].append(f'Reaction {self.arms[i].name}')
                if self.mixing_service.ms:
                    data['component'].append(f'reactant {j+1} ms')
                else:
                    data['component'].append(f'reactant {j+1}')
                data['name'].append(solution.reactant_name)
                data['amount'].append(solution.reactant_amount)
                data['mass'].append(solution.reactant_mass)
                data['volume'].append(solution.reactant_volume)
                data['mix order'].append(j+2+len(combined_solution.base_acid_volumes))
                if self.mixing_service.ms:
                    data['ms-factor'].append(self.mixing_service.reaction.reactants[j].ms_dilution_factor)
                else:
                    data['ms-factor'].append('')
                data['temperature'].append('')
                data['shaker speed'].append('')
                data['time'].append('')

            #catalyst
            for j, catalyst in enumerate(self.service.reaction.catalysts):
                data['experiment name'].append(f'Reaction {self.arms[i].name}')
                data['name'].append(catalyst.name)
                data['component'].append(catalyst.component)
                data['amount'].append(combined_solution.catalyst_amounts[j])
                data['mass'].append(combined_solution.catalyst_masses[j])
                if 'solid' in catalyst.name:
                    data['volume'].append('')
                    data['mix order'].append('')
                else:
                    data['volume'].append(combined_solution.catalyst_volumes[j])
                    data['mix order'].append(j+2+len(combined_solution.solutions)+len(combined_solution.base_acid_volumes))
                data['ms-factor'].append('')
                data['temperature'].append('')
                data['shaker speed'].append('')
                data['time'].append('')
                
            #internal standard
            if self.service.reaction.internal_standard:
                data['experiment name'].append(f'Reaction {self.arms[i].name}')
                data['name'].append(self.service.reaction.internal_standard.name)
                data['component'].append(self.mixing_service.reaction.internal_standard.component)
                data['amount'].append('')
                data['mass'].append('')
                data['volume'].append(combined_solution.internal_standard_volume)
                data['mix order'].append('analysis sample')
                data['ms-factor'].append('')
                data['temperature'].append('')
                data['shaker speed'].append('')
                data['time'].append('')

            #product
            if self.service.reaction.products:
                for j, product in enumerate(self.mixing_service.reaction.products):
                    print('j', j)
                    data['experiment name'].append(f'Reaction {self.arms[i].name}')
                    data['name'].append(product.name+'(quantitative)')
                    data['component'].append(product.component)
                    data['amount'].append(combined_solution.product_amounts[j])
                    data['mass'].append(combined_solution.product_masses[j])
                    data['volume'].append('')
                    data['mix order'].append('')
                    data['ms-factor'].append('')
                    data['temperature'].append('')
                    data['shaker speed'].append('')
                    data['time'].append('')
            
            #analysis_sample
            data['experiment name'].append(f'Reaction {self.arms[i].name}')
            data['component'].append(f"analysis sample {i+1}")
            data['name'].append('')
            data['amount'].append('')
            data['mass'].append('')
            data['volume'].append(self.analysis_volume)
            data['mix order'].append(1)
            data['ms-factor'].append('')
            data['temperature'].append('')
            data['shaker speed'].append('')
            data['time'].append('')
            
            #analysis reaction solution
            data['experiment name'].append(f'Reaction {self.arms[i].name}')
            data['component'].append('reaction solution in analysis sample')
            data['name'].append(f"reaction solution {i+1}")
            data['amount'].append('')
            data['mass'].append('')
            data['volume'].append(self.analysis_volume*(1/self.analysis_dilution_factor))
            data['mix order'].append(1)
            data['ms-factor'].append('')
            data['temperature'].append('')
            data['shaker speed'].append('')
            data['time'].append('')

            #analysis_solvent
            data['experiment name'].append(f'Reaction {self.arms[i].name}')
            data['component'].append('analysis solvent')
            try:
                data['name'].append(self.mixing_service.reaction.analysis_solvent.name)
            except:
                data['name'].append('')
            data['amount'].append('')
            data['mass'].append('')
            data['volume'].append(self.analysis_volume*(1-(1/self.analysis_dilution_factor)))
            data['mix order'].append(1)
            data['ms-factor'].append('')
            data['temperature'].append('')
            data['shaker speed'].append('')
            data['time'].append('')


        df = pd.DataFrame(data)
        df.to_excel(os.path.join(self.data_dir, 'solutions.xlsx'), index=False)

root = tk.Tk()
ui = UI(root)
ui.start_options()

root.mainloop()